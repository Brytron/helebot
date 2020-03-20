import requests
from operator import itemgetter
from openpyxl import load_workbook
from dateutil import parser
from datetime import datetime
from datetime import timedelta
from Cache import CharCache
import aiohttp
import json
import random
import asyncio

bung_url = 'https://www.bungie.net/Platform'
baseurl = 'https://bungie.net/Platform/Destiny2/'
baseurl_groupv2 = 'https://bungie.net/Platform/GroupV2/'

my_api_key = '251bdc1107ac4b4e8713b3fd0aea1257'
my_headers = my_headers = {"X-API-Key": my_api_key}


game_types = { 74:'Control', 72:'Clash', 37:'Survival', 38:'Countdown'}
membership_types = {'xbox': '1',  'xbone': '1', 'psn': '2', 'pc': '4', 'ps4': '2', 'all': '-1'}

D2Profiles = CharCache()
D2Stats = CharCache()

class Season:
    """
    destiny 2 season attaches season start and end date to season
    """
    def __init__(self):
        self.season_id = None
        self.start_time = None
        self.end_time = None

class D2Profile:
    """
    object that contains a users user_id, character_ids, and console_stats
    """
    def __init__(self):
        self.user_id = None
        self.char_ids = None
        self.console = None


class ResponseSummary:
    '''
    Object contains all the important information about the request sent to bungie.
    '''
    def __init__(self, response):
        self.status = response.status_code
        self.url = response.url
        self.data = None
        self.message = None
        self.error_code = None
        self.error_status = None
        self.exception = None
        if self.status == 200:
            result = response.json()
            self.message = result['Message']
            self.error_code = result['ErrorCode']
            self.error_status = result['ErrorStatus']
            if self.error_code == 1:
                try:
                    self.data = result['Response']
                except Exception as ex:
                    print("ResponseSummary: 200 status and error_code 1, but there was no result['Response']")
                    print("Exception: {0}.\nType: {1}".format(ex, ex.__class__.__name__))
                    self.exception = ex.__class__.__name__
            else:
                print('No data returned for url: {0}.\n {1} was the error code with status 200.'.format(self.url, self.error_code))
        else:
            print('Request failed for url: {0}.\n.Status: {0}'.format(self.url, self.status))

    def __repr__(self):
        """What will be displayed/printed for the class instance."""
        disp_header =       "<" + self.__class__.__name__ + " instance>\n"
        disp_data =         ".data: " + str(self.data) + "\n\n"
        disp_url =          ".url: " + str(self.url) + "\n"
        disp_message =      ".message: " + str(self.message) + "\n"
        disp_status =       ".status: " + str(self.status) + "\n"
        disp_error_code =   ".error_code: " + str(self.error_code) + "\n"
        disp_error_status = ".error_status: " + str(self.error_status) + "\n"
        disp_exception =    ".exception: " + str(self.exception)
        return disp_header + disp_data + disp_url + disp_message + \
               disp_status + disp_error_code + disp_error_status + disp_exception


def search_member(search, console):
    """returns name on the specified console"""
    url = bung_url + f"/User/SearchUsers/?q={search}"
    print(url)
    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)

    if console == "pc":
        name = None
        i = 0
        while name is None and i < 20:
            try:
                name = summary.data[i]
            except:
                print(f"searching list[{i}]")
            i = i + 1

    elif console == "xbox":
        name = None
        i = 0
        while name is None and i < 20:
            try:
                name = summary.data[i]['xboxDisplayName']
            except:
                print(f"searching list[{i}]")
            i = i + 1

    elif console == "psn":
        name = None
        i = 0
        while name is None and i < 20:
            try:
                name = summary.data[i]['psnDisplayName']
            except:
                print(f"searching list[{i}]")

    else:
        print("Console not found try lowercase")
        name = None
    return name


def return_user_id(name,console):
    """returns user id from input string"""
    membershipType = membership_types[console]

    if console == "pc":
        blizzard_name = name.split("#")
        name = blizzard_name[0]+ "%23" + blizzard_name[1]

    print(name)
    url = bung_url + f"/Destiny2/SearchDestinyPlayer/{membershipType}/{name}/"
    print(url)
    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)
    return summary.data[0]['membershipId']


def search_user_id(query,console):
    """returns user id(int64) from search querry"""

    membershipType = membership_types[console]
    display_name = search_member(query, console)


    url = bung_url + f"/Destiny2/SearchDestinyPlayer/{membershipType}/{display_name}/"
    print(url)

    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)
    return summary.data[0]['membershipId']


def get_characters(user_id,console):
    """return a list of int64 id player characters"""
    membershipType = membership_types[console]
    destinyMembershipId = user_id
    url = bung_url +  f"/Destiny2/{membershipType}/Profile/{destinyMembershipId}/?components=100"

    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)
    characters = summary.data['profile']['data']['characterIds']
    return characters


def find_profile(name,console):
    """
    Searches a cache for name if in cache return user_id char_id, and char_platform if not in cache search destiny
    api for same information
    """
    key = name + console
    if key in D2Profiles.cache:
        print("user in cache")
        return D2Profiles.cache[key]['value']
    else:
        profile = D2Profile()

        profile.user_id = return_user_id(name, console)

        profile.char_ids = get_characters(profile.user_id, console)
        profile.console = console
        D2Profiles.update(key, profile)
        print("user added from Bungie API")
        return profile


def get_stats(user_id,console):
    """gets overall stats for character"""
    membershipType = membership_types[console]
    destinyMembershipId = user_id
    url = bung_url + f"/Destiny2/{membershipType}/Account/{destinyMembershipId}/Stats/"
    print(url)

    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)
    return summary.data


def find_stats(profile, console):
    """gets overall stats from cache or api"""

    key = profile.user_id + console
    if key in D2Stats.cache and D2Stats.cache[key]['date_accessed'] > datetime.now() - timedelta(days=1):
        print(D2Stats.cache[key]['date_accessed'])
        print("stats in cache")
        return D2Stats.cache[key]['value']
    else:
        stat = get_stats(profile.user_id, console)
        D2Stats.update(key, stat)
        print("stats added from Bungie API")
        return stat


def get_historical_stats(destinyMembershipId,console,characterId,daystart,dayend,modes):
    """gets historical stats for daystart to dayend"""
    membershipType = membership_types[console]
    url = bung_url + f'/Destiny2/{membershipType}/Account/{destinyMembershipId}/Character/{characterId}/Stats/' \
    + f'?daystart={daystart}&dayend={dayend}&modes={modes}'
    print(url)

    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)
    return summary.data


def get_historical_activitys(destinyMembershipId,console,characterId,count,page = '0',mode = '0'):
    """gets historical stats for daystart to dayend"""
    membershipType = membership_types[console]
    url = bung_url + f'/Destiny2/{membershipType}/Account/{destinyMembershipId}/Character/{characterId}/Stats/Activities/' \
    + f'?count={count}&mode={mode}&page={page}'
    print(url)

    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)
    print(summary.error_status)
    return summary.data


def build_comp(name, console, season):
    """returns stats in a list for comp in a season"""
    if season == "S7":
        season_start = "2019-06-04T17:00:00Z"
        season_end = "2019-10-01T17:00:00Z"
    elif season == "S6":
        season_start = "2019-03-05T17:00:00Z"
        season_end = "2019-06-04T17:00:00Z"
    else:
        print("this season is not currently supported")
        return

    profile = find_profile(name,console)
    win_list = []
    for character in profile.char_ids:
        page = 0
        try:
            stats = get_historical_activitys(profile.user_id, console, character, '5', '0', '69')
            while 'activities' in stats and stats['activities'][0]['period']>season_start:
                page = page + 1
                num = 0
                while num < 5 and stats['activities'][num]['period']>season_start and stats['activities'][num]['period']<season_end:
                    win_dict = {}
                    win_dict['period'] = stats['activities'][num]['period']
                    win_dict['win/loss'] = stats['activities'][num]['values']['standing']['basic']['displayValue']
                    win_dict['kills'] = stats['activities'][num]['values']['kills']['basic']['value']
                    win_dict['assists'] = stats['activities'][num]['values']['assists']['basic']['value']
                    win_dict['deaths'] = stats['activities'][num]['values']['deaths']['basic']['value']
                    win_dict['mode'] = stats['activities'][num]['activityDetails']['mode']
                    win_list.append(win_dict)
                    print(win_dict)
                    num = num + 1
                stats = get_historical_activitys(profile.user_id, console, character, '5', str(page), '69')
        except TypeError:
            print("comp history does not exist for this character")

    return sorted(win_list, key=itemgetter('period'))


def pushto_excel(comp_build, season, savename):
    """push win and loss data to excel sheet"""
    if season == "S7":
        season_start = "2019-06-04T17:00:00Z"
        season_end = "2019-09-17T17:00:00Z"
    elif season == "S6":
        season_start = "2019-03-05T17:00:00Z"
        season_end = "2019-06-04T17:00:00Z"
    else:
        print("this season is not currently supported")
        return

    season_start_date = parser.parse(season_start)
    season_end_date = parser.parse(season_end)
    next_reset = season_start_date + timedelta(days=7)
    print(next_reset)

    wb = load_workbook(filename = 'glory_base.xlsx')
    ws = wb.active
    ws['G22'] = season_start_date.date()
    ws['H22'] = season_end_date.date()
    col = 32
    game_col = 32
    for game in comp_build:
        ws['R'+str(game_col)] = game_col - 31
        current_cell = 'O' + str(col)
        print(current_cell)
        game_date = parser.parse(game['period'])
        while game_date > next_reset:
            ws[current_cell] = "G"
            col = col + 1
            current_cell = 'O' + str(col)
            next_reset = next_reset + timedelta(days=7)
        if game['win/loss'] == "Victory":
            ws[current_cell] = "W"
            ws['S'+str(game_col)] = "W"
        else:
            ws[current_cell] = "L"
            ws['S' + str(game_col)] = "L"
        ws['T' + str(game_col)] = game['kills']
        ws['U' + str(game_col)] = game['assists']
        ws['V' + str(game_col)] = game['deaths']
        try:
            ws['AA' + str(game_col)] = game_types[game['mode']]
        except KeyError:
            ws['AA' + str(game_col)] = game['mode']
        ws['W' + str(game_col)] = game_date.date()
        ws['X' + str(game_col)] = game_date.time()
        col = col + 1
        game_col = game_col+1
    filename = f'{season}glory_{savename}.xlsx'
    print(filename)
    wb.save(filename)
    return filename


def save_historical_stats():
    user_id = return_user_id("Congohunter#11767","pc")
    stats = get_stats(user_id,"pc")
    out = json.dumps(stats)
    f = open("stats.json","w")
    f.write(out)
    f.close


def view_stats():
    with open ('stats.json') as stats_file:
        stats = json.load(stats_file)

    one_dict = stats["mergedAllCharacters"]['results']['allPvE']['allTime']
    random_key = random.sample(list(one_dict),1)[0]
    print(random_key)
    value = one_dict[random_key]["basic"]["value"]
    print(value)
    random_key = random_key[:1].upper() + random_key[1:]
    print(random_key)
    pos = [i for i, e in enumerate(random_key + 'A') if e.isupper()]
    parts = [random_key[pos[j]:pos[j + 1]] for j in range(len(pos) - 1)]
    return parts


def cut_stat_key(key):
    key = key[:1].upper() + key[1:]
    pos = [i for i, e in enumerate(key + 'A') if e.isupper()]
    parts = [key[pos[j]:pos[j + 1]] for j in range(len(pos) - 1)]
    return parts


def remove_hashtag(name):
    return name.split("#")[0]


def fight_stats(player_1,player_2,console):
    if console == "pc":
        name_1 = remove_hashtag(player_1)
        name_2 = remove_hashtag(player_2)
    else:
        name_1 = player_1
        name_2 = player_2

    profile_1 = find_profile(player_1, console)

    stats_1 = find_stats(profile_1, console)

    enemy_type = random.choice(['allPvP', 'allPvE'])
    dict_1 = stats_1["mergedAllCharacters"]['results'][enemy_type]['allTime']

    random_key = random.sample(list(dict_1), 1)[0]
    print(random_key)
    cut_key = cut_stat_key(random_key)

    profile_2 = find_profile(player_2,console)

    stats_2 = find_stats(profile_2, console)
    dict_2 = stats_2["mergedAllCharacters"]['results'][enemy_type]['allTime']

    value_1 = dict_1[random_key]["basic"]["value"]
    value_2 = dict_2[random_key]["basic"]["value"]
    val_dif = abs(value_1-value_2)
    if val_dif > 10:
        val_dif = "{:,.0f}".format(val_dif)
    else:
        val_dif = "{:,.2f}".format(val_dif)

    if value_1 > value_2:
        s = f"{name_1} has {val_dif} more"
        for word in cut_key:
            s = s + f" {word.lower()}"
        s = s + f" than {name_2}"
    elif value_2 > value_1:
        s = f"{name_2} has {val_dif} more"
        for word in cut_key:
            s = s + f" {word.lower()}"
        s = s + f" than {name_1}"
    else:
        s = f"Both players have {value_1}"
        for word in cut_key:
            s = s + f" {word.lower()}"
    s = s + f" in {enemy_type[3:]}"
    return s

"""
def get_linked_profiles(membershipId):

    url =  bung_url + f"/Destiny2/-1/Profile/{membershipId}/LinkedProfiles/"
    response = requests.get(url, headers=my_headers)
    summary = ResponseSummary(response)
    return summary
"""

#def test_comp():
    #name = search_member("Brytron","pc")
    #print(name)
    #user_id = find_profile("Brytron#11867","pc")
    #print(user_id)
    #user_id = find_profile("Congohunter#11767","pc")
    #print(user_id)
    #characters = get_characters(user_id,"pc")
    #print(characters)
    #user_id = search_user_id("Brytron","pc")
    #print(user_id)
    #user_id = return_user_id("Brytron#11867","all")
    #print(user_id)
    #print(return_user_id("Brytron#11867", "pc"))
    #print(fight_stats("Brytron#11867", "Congohunter#11767", "pc"))
    #print(fight_stats("Brytron#11867", "Congohunter#11767", "pc"))
    #print(fight_stats("Brytron#11867", "Congohunter#11767", "pc"))
    #print(fight_stats("Brytron#11867", "Congohunter#11767", "pc"))
    #print(fight_stats("Brytron#11867", "Congohunter#11767", "pc"))

#test_comp()

