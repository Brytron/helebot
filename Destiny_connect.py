import aiohttp
import asyncio
import random
import pickle
import os.path
import json
import pandas as pd

from Cache import CharCache
from operator import itemgetter
from openpyxl import load_workbook
from dateutil import parser
from datetime import datetime
from datetime import timedelta
from pytz import timezone


class D2Profile:
    """
    object that contains a users user_id, character_ids, and console_stats
    """
    def __init__(self):
        self.title = None
        self.user_id = None
        self.char_ids = None
        self.console = None

reset_time_hours = 17
bung_url = 'https://www.bungie.net/Platform'
baseurl = 'https://bungie.net/Platform/Destiny2/'
baseurl_groupv2 = 'https://bungie.net/Platform/GroupV2/'

with open("config.json", "r") as f:
    my_api_key = json.load(f)["Destiny"]["APICode"]

my_headers = {"X-API-Key": my_api_key}
membership_types = {'xbox': '1',  'xbone': '1', 'psn': '2', 'steam': '3', 'ps4': '2', 'all': '-1', 'blizzard': '4',
                    'demon': '10', 'next': '254', 'stadia': '5'}
seasons = \
    {'S7': ["2019-06-04T17:00:00Z", "2019-10-01T17:00:00Z"], 'S6': ["2019-03-05T17:00:00Z", "2019-06-04T17:00:00Z"],
     'S8': ["2019-10-01T17:00:00Z", "2019-12-010T17:00:00Z"], 'S9': ["2019-12-10T17:00:00Z", "2020-03-09T17:00:00Z"],
     'S10': ["2020-03-09T17:00:00Z", "2020-06-09T17:00:00Z"], 'S11': ["2020-06-09T17:00:00Z", "2020-11-10T17:00:00Z"],
     'S12': ["2020-11-10T17:00:00Z", "2021-03-10T17:00:00Z"], 'S13': ["2021-03-10T17:00:00Z", "2021-05-11T17:00:00Z"],
     'S14': ["2021-05-11T17:00:00Z", "2021-08-11T17:00:00Z"], 'S15': ["2021-08-24T17:00:00Z", "2022-02-22T17:00:00Z"]}
current_season = 'S15'
game_types = { 74:'Control', 72:'Clash', 37:'Survival', 38:'Countdown'}
trials_game_type = {84:'Elimination'}

#D2Profiles = CharCache()
D2Stats = CharCache()

def replace_hash(name_in):
    "replaces the '#' in user_input of bungie names"
    name_api = name_in.replace("#", "%23")
    return name_api

def load_profiles():
    """returns  in d2 character profiles from a pickled file"""
    Profiles = None
    if os.path.exists('d2CharCache.pickle'):
        with open('d2CharCache.pickle', 'rb') as Char_file:
            Profiles = pickle.load(Char_file)
    if not Profiles:
        Profiles = CharCache()
        with open('d2CharCache.pickle', 'wb') as char_file:
            pickle.dump(Profiles, char_file)
    return Profiles

D2Profiles = load_profiles()

async def fetch(session, url):
    async with session.request('GET',url,headers=my_headers) as response:
        return await response.json()



async def get_characters(session, membership_id, console):
    """return a list of int64 id player characters"""

    membershipType = membership_types[console]
    print(membershipType)
    destinyMembershipId = membership_id
    url = bung_url + f"/Destiny2/{membershipType}/Profile/{destinyMembershipId}/?components=100"

    response = await fetch(session, url)
    print(response)
    characters = response['Response']['profile']['data']['characterIds']
    return characters


async def return_member_id(session, name, console):
    """returns user id from input string"""
    search_membershipType = membership_types["all"]
    user_selected_membershipType = membership_types[console]

    print(name)
    name = replace_hash(name)
    url = bung_url + f"/Destiny2/SearchDestinyPlayer/{search_membershipType}/{name}/"
    print(url)
    response = await fetch(session, url)
    membership_id = None
    for item in response['Response']:
        print(item)
        print(user_selected_membershipType)
        print(item['applicableMembershipTypes'])
        print(item['membershipId'])
        for i in item['applicableMembershipTypes']:
            if str(i) == str(user_selected_membershipType):
                print('member id found')
                return item['membershipId']

    return membership_id



async def return_member_id_from_code(session, credential, console):
    """returns a id from a steam code, this can probably be removed at some point"""
    membershipType = membership_types[console]

    crType = 'SteamId'
    url = bung_url + f"/User/GetMembershipFromHardLinkedCredential/{crType}/{credential}/"
    print(url)
    response = await fetch(session, url)

    return response['Response']['membershipId']


async def register(title, bungieID, console):
    """
    connects a users name and bungieID and create a profile for a user
    """
    async with aiohttp.ClientSession() as session:
        profile = D2Profile()
        profile.user_id = await return_member_id(session, bungieID, console)
        profile.char_ids = await get_characters(session, profile.user_id, console)
        profile.console = console
        key = title + console
        D2Profiles.update(key, profile)
        with open('d2CharCache.pickle', 'wb') as char_file:
            pickle.dump(D2Profiles, char_file)
        print("registration complete")
        if key in D2Profiles.cache:
            print("registration success")
            error = 1
        else:
            print("registration failure")
            error = 0
        return error


async def find_profile(session, name, console):
    """
    Searches a cache for name if in cache return user_id char_id, and char_platform if not in cache search destiny
    api for same information
    """
    key = name + console
    if key in D2Profiles.cache:
        print("user in cache")
        return D2Profiles.cache[key]['value']
    else:
        profile = D2Profile()
        profile.user_id = await return_member_id(session, name, console)
        profile.char_ids = await get_characters(session, profile.user_id, console)
        profile.console = console
        D2Profiles.update(key, profile)
        with open('d2CharCache.pickle', 'wb') as char_file:
            pickle.dump(D2Profiles, char_file)
        print("user added from Bungie API")
        return profile


async def get_historical_activitys(session,destinyMembershipId,console,characterId,count,page = '0',mode = '0'):
    """gets historical stats for daystart to dayend"""
    membershipType = membership_types[console]
    url = bung_url + f'/Destiny2/{membershipType}/Account/{destinyMembershipId}/Character/{characterId}/Stats/Activities/' \
    + f'?count={count}&mode={mode}&page={page}'
    print(url)

    response = await fetch(session, url)

    return response['Response']


def find_scope_time(season=None,week=None):

    if season is None:
        if week is None:
            if datetime.utcnow().hour > reset_time_hours:
                today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)-timedelta(days=1)
        else:
            today = parser.parse(week)

        if today.weekday() == 1:
            offset = 0
        else:
            offset = (today.weekday() - 1) % 7

        reset = today - timedelta(days=offset) + timedelta(hours=reset_time_hours)
        reset_next = reset + timedelta(days=7)

        search_on = {'start':str(reset.isoformat() + "Z"),'stop':str(reset_next.isoformat() + "Z")}

    season_times = seasons[season]
    search_on = {'start': season_times[0], 'stop': season_times[1]}

    return search_on


async def build_banner(session, name, console, start_search_on, stop_search_on):

    profile = await find_profile(session, name, console)
    win_list = []
    for character in profile.char_ids:
        page = 0
        try:
            stats = await get_historical_activitys(session, profile.user_id, console, character, '5', '0', '19')
            print(stats)
            while 'activities' in stats and stats['activities'][0]['period'] > start_search_on:
                page = page + 1
                num = 0
                while num < 5 and stats['activities'][num]['period'] > start_search_on and stats['activities'][num]['period'] < stop_search_on:
                    win_dict = {}
                    win_dict['period'] = stats['activities'][num]['period']
                    win_dict['win/loss'] = stats['activities'][num]['values']['standing']['basic']['displayValue']
                    win_dict['kills'] = stats['activities'][num]['values']['kills']['basic']['value']
                    win_dict['assists'] = stats['activities'][num]['values']['assists']['basic']['value']
                    win_dict['deaths'] = stats['activities'][num]['values']['deaths']['basic']['value']
                    win_dict['mode'] = stats['activities'][num]['activityDetails']['mode']
                    win_dict['duration'] = stats['activities'][num]['values']['activityDurationSeconds']['basic']['value']
                    win_list.append(win_dict)
                    print(win_dict)
                    num = num + 1
                stats = await get_historical_activitys(session, profile.user_id, console, character, '5', str(page), '19')

        except TypeError:
            print("banner history does not exist for this character")

    return sorted(win_list, key=itemgetter('period'))


def pushto_excel_banner(banner_build, time_start, time_stop, savename, tz):
    """push win and loss data to excel sheet"""
    time_start = parser.parse(time_start).astimezone(timezone(tz))
    time_stop = parser.parse(time_stop).astimezone(timezone(tz))
    wb = load_workbook(filename='banner_base.xlsx')
    ws = wb.active
    ws['B2'] = time_start.date()
    ws['B3'] = time_stop.date()
    game_col = 6
    for game in banner_build:
        ws['A'+str(game_col)] = game_col - 5
        game_date = parser.parse(game['period']).astimezone(timezone(tz))

        if game['win/loss'] == "Victory":
            ws['B'+str(game_col)] = "W"
        else:
            ws['B' + str(game_col)] = "L"

        ws['C' + str(game_col)] = game['kills']
        ws['D' + str(game_col)] = game['assists']
        ws['E' + str(game_col)] = game['deaths']
        ws['H' + str(game_col)] = game_date.date()
        ws['I' + str(game_col)] = game_date.time()
        ws['N' + str(game_col)] = game['duration']
        game_col = game_col+1

    filename = f'{time_start.strftime("%m_%d_%y")}banner_{savename}.xlsx'
    wb.save(filename)
    return filename


async def banner_stats(name, console="pc", season=None, week=None, tz="US/Pacific"):
    async with aiohttp.ClientSession() as session:
        time = find_scope_time(season,week)
        build = await build_banner(session, name, console, time['start'], time['stop'])
        return pushto_excel_banner(build, time['start'], time['stop'], name, tz)


async def build_trials(session, name, console, start_search_on, stop_search_on):

    profile = await find_profile(session, name, console)
    win_list = []
    for character in profile.char_ids:
        page = 0
        try:
            stats = await get_historical_activitys(session, profile.user_id, console, character, '5', '0', '84')
            print(stats)
            while 'activities' in stats and stats['activities'][0]['period'] > start_search_on:
                page = page + 1
                num = 0
                next_time = stats['activities'][num]['period']
                while num < 5 and next_time > start_search_on and stats['activities'][num]['period'] < stop_search_on:
                    win_dict = {}
                    win_dict['period'] = stats['activities'][num]['period']
                    win_dict['win/loss'] = stats['activities'][num]['values']['standing']['basic']['displayValue']
                    win_dict['kills'] = stats['activities'][num]['values']['kills']['basic']['value']
                    win_dict['assists'] = stats['activities'][num]['values']['assists']['basic']['value']
                    win_dict['deaths'] = stats['activities'][num]['values']['deaths']['basic']['value']
                    win_dict['mode'] = stats['activities'][num]['activityDetails']['mode']
                    win_dict['duration'] = stats['activities'][num]['values']['activityDurationSeconds']['basic']['value']
                    win_list.append(win_dict)
                    print(win_dict)
                    num = num + 1
                    try:
                        next_time = stats['activities'][num]['period']
                    except IndexError:
                        next_time = 0
                        print("end of games")

                stats = await get_historical_activitys(session, profile.user_id, console, character, '5', str(page), '84')

        except TypeError:
            print("trials history does not exist for this character")

    return sorted(win_list, key=itemgetter('period'))


def create_trials_df(user_name, in_dict):
    "returns a dataframe with username appended to column fields"
    return pd.DataFrame(data=in_dict).set_index('period').rename(
        columns={"win/loss": user_name + " win/loss", "kills": user_name + " kills", "assists": user_name + " assists",
                 "deaths": user_name + " deaths", "mode": user_name + " mode", "duration": user_name + " duration"}
    )


def pushto_excel_trials_team(in_df, time_start, time_stop, savename, tz, user1, user2, user3):
    """push win and loss data to excel sheet note full trial_dict in for loop is unoptimised"""
    time_start = parser.parse(time_start).astimezone(timezone(tz))
    time_stop = parser.parse(time_stop).astimezone(timezone(tz))
    wb = load_workbook(filename='trialsteam_base.xlsx')
    ws = wb.active
    trials_dict = in_df.T.to_dict()
    print(trials_dict)
    ws['B2'] = time_start.date()
    ws['B3'] = time_stop.date()
    ws['C4'] = user1
    ws['H4'] = user2
    ws['M4'] = user3
    game_col = 6
    for game in trials_dict:
        ws['A'+str(game_col)] = game_col - 5
        print(game)
        game_date = parser.parse(game).astimezone(timezone(tz))

        if trials_dict[game][user1 + ' win/loss'] == "Victory":
            ws['B'+str(game_col)] = "W"
        else:
            ws['B'+str(game_col)] = "L"

        ws['C' + str(game_col)] = trials_dict[game][user1+' kills']
        ws['D' + str(game_col)] = trials_dict[game][user1+' assists']
        ws['E' + str(game_col)] = trials_dict[game][user1+' deaths']

        ws['H' + str(game_col)] = trials_dict[game][user2+' kills']
        ws['I' + str(game_col)] = trials_dict[game][user2+' assists']
        ws['J' + str(game_col)] = trials_dict[game][user2+' deaths']

        if user3 is not None:
            ws['M' + str(game_col)] = trials_dict[game][user3 + ' kills']
            ws['N' + str(game_col)] = trials_dict[game][user3 + ' assists']
            ws['O' + str(game_col)] = trials_dict[game][user3 + ' deaths']

        ws['T' + str(game_col)] = game_date.date()
        ws['U' + str(game_col)] = game_date.time()
        ws['Z' + str(game_col)] = trials_dict[game][user1+' duration']
        game_col = game_col+1

    filename = f'{time_start.strftime("%m_%d_%y")}trialsteam_{savename}.xlsx'
    wb.save(filename)
    return filename


async def build_trials_team(user1, user2, user3=None, console='steam', season=current_season, week=None, tz="US/Pacific"):
    "using 2 or more player create a collection of stats of games where the user played together"
    async with aiohttp.ClientSession() as session:
        out_df = None
        time = find_scope_time(season, week)
        user1_games = await build_trials(session, user1, console, time['start'], time['stop'])
        user2_games = await build_trials(session, user2, console, time['start'], time['stop'])
        user1_df = create_trials_df(user1, user1_games)
        user2_df = create_trials_df(user2, user2_games)
        print(user1_df)
        out_df = user1_df.join(user2_df, on='period', how='inner')
        if user3 is not None:
            user3_games = await build_trials(session, user3, console, time['start'], time['stop'])
            user3_df = create_trials_df(user3, user3_games)
            out_df = out_df.join(user3_df, on='period', how='inner')
        filename = pushto_excel_trials_team(
            out_df, time['start'], time['stop'], user1+user2+str(user3), tz, user1, user2, user3)
        print(filename)
        return filename


def pushto_excel_trials(trials_build, time_start, time_stop, savename, tz):
    """push win and loss data to excel sheet"""
    time_start = parser.parse(time_start).astimezone(timezone(tz))
    time_stop = parser.parse(time_stop).astimezone(timezone(tz))
    wb = load_workbook(filename='trials_base.xlsx')
    ws = wb.active
    ws['B2'] = time_start.date()
    ws['B3'] = time_stop.date()
    game_col = 6
    for game in trials_build:
        ws['A'+str(game_col)] = game_col - 5
        game_date = parser.parse(game['period']).astimezone(timezone(tz))

        if game['win/loss'] == "Victory":
            ws['B'+str(game_col)] = "W"
        else:
            ws['B' + str(game_col)] = "L"

        ws['C' + str(game_col)] = game['kills']
        ws['D' + str(game_col)] = game['assists']
        ws['E' + str(game_col)] = game['deaths']
        ws['H' + str(game_col)] = game_date.date()
        ws['I' + str(game_col)] = game_date.time()
        ws['N' + str(game_col)] = game['duration']
        game_col = game_col+1

    filename = f'{time_start.strftime("%m_%d_%y")}trials_{savename}.xlsx'
    wb.save(filename)
    return filename


async def trials_stats(name, console="pc", season=None, week=None, tz="US/Pacific"):
    async with aiohttp.ClientSession() as session:
        time = find_scope_time(season,week)
        build = await build_trials(session, name, console, time['start'], time['stop'])
        return pushto_excel_trials(build, time['start'], time['stop'], name, tz)


async def build_comp(session, name, console, season):
    """returns stats in a list for comp in a season"""

    season_times = seasons[season]
    season_start = season_times[0]
    season_end = season_times[1]

    profile = await find_profile(session, name, console)
    win_list = []
    for character in profile.char_ids:
        page = 0
        try:
            stats = await get_historical_activitys(session, profile.user_id, console, character, '5', '0', '69')
            while 'activities' in stats and stats['activities'][0]['period'] > season_start:
                page = page + 1
                num = 0
                while num < 5 and stats['activities'][num]['period'] > season_start and stats['activities'][num]['period'] < season_end:
                    win_dict = {}
                    win_dict['period'] = stats['activities'][num]['period']
                    win_dict['win/loss'] = stats['activities'][num]['values']['standing']['basic']['displayValue']
                    win_dict['kills'] = stats['activities'][num]['values']['kills']['basic']['value']
                    win_dict['assists'] = stats['activities'][num]['values']['assists']['basic']['value']
                    win_dict['deaths'] = stats['activities'][num]['values']['deaths']['basic']['value']
                    win_dict['mode'] = stats['activities'][num]['activityDetails']['mode']
                    win_list.append(win_dict)
                    print(win_dict)
                    num = num + 1
                stats = await get_historical_activitys(session, profile.user_id, console, character, '5', str(page), '69')
        except TypeError:
            print("comp history does not exist for this character")

    return sorted(win_list, key=itemgetter('period'))


def pushto_excel_comp(comp_build, season, savename, tz):
    """push win and loss data to excel sheet"""

    season_times = seasons[season]
    season_start = season_times[0]
    season_end = season_times[1]

    season_start_date = parser.parse(season_start).astimezone(timezone(tz))
    season_end_date = parser.parse(season_end).astimezone(timezone(tz))
    next_reset = season_start_date + timedelta(days=7)
    print(next_reset)

    wb = load_workbook(filename='glory_base.xlsx')
    ws = wb.active
    ws['G22'] = season_start_date.date()
    ws['H22'] = season_end_date.date()
    col = 32
    game_col = 32
    for game in comp_build:
        ws['R'+str(game_col)] = game_col - 31
        current_cell = 'O' + str(col)
        print(current_cell)
        game_date = parser.parse(game['period']).astimezone(timezone(tz))
        while game_date > next_reset:
            ws[current_cell] = "G"
            col = col + 1
            current_cell = 'O' + str(col)
            next_reset = next_reset + timedelta(days=7)
        if game['win/loss'] == "Victory":
            ws[current_cell] = "W"
            ws['S'+str(game_col)] = "W"
        else:
            ws[current_cell] = "L"
            ws['S' + str(game_col)] = "L"
        ws['T' + str(game_col)] = game['kills']
        ws['U' + str(game_col)] = game['assists']
        ws['V' + str(game_col)] = game['deaths']
        try:
            ws['AA' + str(game_col)] = game_types[game['mode']]
        except KeyError:
            ws['AA' + str(game_col)] = game['mode']
        ws['W' + str(game_col)] = game_date.date()
        ws['X' + str(game_col)] = game_date.time()
        col = col + 1
        game_col = game_col+1
    filename = f'{season}glory_{savename}.xlsx'
    print(filename)
    wb.save(filename)
    return filename


async def competitive_stats(name,console,season):
    async with aiohttp.ClientSession() as session:
        return await build_comp(session,name,console,season)


async def get_stats(session, user_id, console):
    """gets overall stats for character"""
    membershipType = membership_types[console]
    destinyMembershipId = user_id
    url = bung_url + f"/Destiny2/{membershipType}/Account/{destinyMembershipId}/Stats/"
    print(url)

    response = await fetch(session, url)
    return response['Response']


async def find_stats(session, profile, console):
    """gets overall stats from cache or api"""

    key = profile.user_id + console
    if key in D2Stats.cache and D2Stats.cache[key]['date_accessed'] > datetime.now() - timedelta(days=1):
        print(D2Stats.cache[key]['date_accessed'])
        print("stats in cache")
        return D2Stats.cache[key]['value']
    else:
        stat = await get_stats(session, profile.user_id, console)
        D2Stats.update(key, stat)
        print("stats added from Bungie API")
        return stat


def cut_stat_key(key):
    key = key[:1].upper() + key[1:]
    pos = [i for i, e in enumerate(key + 'A') if e.isupper()]
    parts = [key[pos[j]:pos[j + 1]] for j in range(len(pos) - 1)]
    return parts


def remove_hashtag(name):
    return name.split("#")[0]


async def fight_stats(player_1,player_2,console):
    async with aiohttp.ClientSession() as session:
        if console == "pc":
            name_1 = remove_hashtag(player_1)
            name_2 = remove_hashtag(player_2)
        else:
            name_1 = player_1
            name_2 = player_2

        profile_1 = await find_profile(session, player_1, console)

        stats_1 = await find_stats(session, profile_1, console)

        enemy_type = random.choice(['allPvP', 'allPvE'])
        dict_1 = stats_1["mergedAllCharacters"]['results'][enemy_type]['allTime']

        random_key = random.sample(list(dict_1), 1)[0]
        print(random_key)
        cut_key = cut_stat_key(random_key)

        profile_2 = await find_profile(session, player_2, console)

        stats_2 = await find_stats(session, profile_2, console)
        dict_2 = stats_2["mergedAllCharacters"]['results'][enemy_type]['allTime']

        value_1 = dict_1[random_key]["basic"]["value"]
        value_2 = dict_2[random_key]["basic"]["value"]
        val_dif = abs(value_1-value_2)
        if val_dif > 10:
            val_dif = "{:,.0f}".format(val_dif)
        else:
            val_dif = "{:,.2f}".format(val_dif)

        if value_1 > value_2:
            s = f"{name_1} has {val_dif} more"
            for word in cut_key:
                s = s + f" {word.lower()}"
            s = s + f" than {name_2}"
        elif value_2 > value_1:
            s = f"{name_2} has {val_dif} more"
            for word in cut_key:
                s = s + f" {word.lower()}"
            s = s + f" than {name_1}"
        else:
            s = f"Both players have {value_1}"
            for word in cut_key:
                s = s + f" {word.lower()}"
        s = s + f" in {enemy_type[3:]}"
        return s


async def example():
    #test = await banner_stats("Brytron","steam",season="S11")

    async with aiohttp.ClientSession() as session:
        membershipType = membership_types['all']
        # membershipType = membership_types[console]
        name2 = 'Congohunter#2140'
        name1 = 'Bryton#3746'
        name3 = 'Jet3010#9562'
        await build_trials_team(name1, name2, name3)
        #print(response['Response'][0]['membershipId'])
        #test = await get_characters(session, name, "all")
        #print(test)


if __name__ == '__main__':
    loop = asyncio.get_event_loop()
    loop.run_until_complete(example())

